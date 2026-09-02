"""Independent check on client-reporting-dashboard.xlsx.

openpyxl writes formulas but never evaluates them, so a wrong SUMIFS would sit
in the file looking perfectly fine. This recomputes every rollup in plain Python
straight from the Data tab and prints what the workbook must show when opened.

Run:  uv run --with openpyxl python verify-dashboard.py
Open the .xlsx (or the Google Sheets import) and compare. Any mismatch is a bug
in the workbook, not in this script.
"""
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

XLSX = Path(__file__).with_name("client-reporting-dashboard.xlsx")

wb = openpyxl.load_workbook(XLSX)          # formulas as written
data = wb["Data"]
calc = wb["Calc"]

# ---- structural assertions -------------------------------------------------
assert wb.sheetnames == ["How this works", "Data", "Calc", "Dashboard"], wb.sheetnames
assert [c.value for c in data[1]] == [
    "Date", "Category", "Vendor", "Description", "Amount (USD)"
], "Data header changed - Calc formulas assume this column order"

# The whole claim of the sample is that Calc never hardcodes a row count.
bad = []
for row in calc.iter_rows():
    for cell in row:
        if isinstance(cell.value, str) and cell.value.startswith("=") and "Data!" in cell.value:
            if "Data!$E:$E" not in cell.value and "Data!$A:$A" not in cell.value \
               and "Data!$B:$B" not in cell.value:
                bad.append((cell.coordinate, cell.value))
assert not bad, f"Calc references Data by a bounded range - re-import will break: {bad}"

# ---- recompute the rollups independently -----------------------------------
actual = defaultdict(float)                # (category, month) -> total
per_cat = defaultdict(float)
total = 0.0
n = 0
for d, cat, _vendor, _desc, amt in data.iter_rows(min_row=2, values_only=True):
    if d is None:
        continue
    actual[(cat, (d.year, d.month))] += amt
    per_cat[cat] += amt
    total += amt
    n += 1

budget_monthly = {}
for r in range(18, 26):                    # budget block, per build script
    cat = calc.cell(row=r, column=1).value
    budget_monthly[cat] = calc.cell(row=r, column=2).value

MONTHS = [(2026, m) for m in range(1, 7)]
n_months = len(MONTHS)

print(f"{XLSX.name} - {n} transactions\n")
print(f"{'Category':<28}{'Budget YTD':>12}{'Actual YTD':>12}{'Variance':>12}  Status")
print("-" * 78)

over = 0
budget_total = 0.0
for cat, monthly in budget_monthly.items():
    b = monthly * n_months
    a = per_cat[cat]
    v = a - b
    budget_total += b
    status = "OVER" if v > 0 else "Within budget"
    if v > 0:
        over += 1
    print(f"{cat:<28}{b:>12,.0f}{a:>12,.0f}{v:>+12,.0f}  {status}")

print("-" * 78)
print(f"{'TOTAL':<28}{budget_total:>12,.0f}{total:>12,.0f}{total - budget_total:>+12,.0f}")
print()
print("Dashboard KPI tiles should read:")
print(f"  TOTAL SPEND YTD   ${total:,.0f}")
print(f"  TOTAL BUDGET YTD  ${budget_total:,.0f}")
print(f"  VARIANCE          ${total - budget_total:+,.0f}")
print(f"  CATEGORIES OVER   {over}")
print()
print("Monthly totals (the trend chart):")
for y, m in MONTHS:
    mt = sum(v for (c, ym), v in actual.items() if ym == (y, m))
    print(f"  {y}-{m:02d}  ${mt:,.0f}")

# The sample is worthless as a talking point if nothing is ever flagged.
assert 1 <= over <= 4, (
    f"{over} of 8 categories over budget. If everything is red nothing is flagged, "
    "and if nothing is red the conditional formatting demonstrates nothing.")
assert n > 100, f"only {n} transactions; sample should look like real volume"
print(f"\nOK - structure, whole-column refs, and {over} flagged overruns all check out.")
